"""

Allows a user to investigate a run's data matrix, displays area/height matrices and colors outlier
cells red.  Supplies which metric the peak is an outlier with respect to, as well as color codes
samples according to group and colors the standards for easy viewing.  Includes metrics of per-sample
and per-molecule metrics for easy calculations.

left click a cell to see the metrics with which it is an outlier in respect to, right click allows the 
user to view the chromatogram which navigates to the peak view for the selected peak in the chromatogram
tab.

Also allows the user to export the data to excel, which exports the data selected that is displayed (area, heihgt,
any qc metric) to be exported to an excel file, keeping color coding and including a metadat tab that has
sample and molecule tables from the sql database as well as an outliers tab showing sample, molecule, list of
metrics with which the sample is an outlier with respect to.

"""

# region Imports

import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from PyQt5.QtWidgets import (QWidget, QTableWidget, QVBoxLayout, QHBoxLayout, QSplitter,
                             QLabel, QComboBox, QHeaderView, QSizePolicy, QTableWidgetItem,
                             QGroupBox, QStyledItemDelegate, QFileDialog, QPushButton, QMenu,
                             QMessageBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor, QBrush

import matplotlib.cm as cm

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    filename="debug.log"
)
logger = logging.getLogger(__name__)

# endregion

# Main data table
class DataTab(QWidget):
    def __init__(self, run_data, chrom_tab, parent):
        super().__init__(parent)

        # save relevant run data
        self.chrom_tab = chrom_tab
        self.run_data = run_data
        self.data_matrix = run_data.data_matrix
        self.n_samples = self.data_matrix.n_samples
        self.sample_map = self.data_matrix.sample_map
        self.inv_sample_map = {v:k for k,v in self.sample_map.items()}
        self.n_molecules = self.data_matrix.n_molecules
        self.mol_map = self.data_matrix.mol_map
        self.inv_mol_map = {v:k for k,v in self.mol_map.items()}
        self.data_metrics = list(self.data_matrix.data.keys())
        self.group_map = self.data_matrix.group_map
        self.group_indices = self.data_matrix.group_indices

        # group colormapping
        groups = list(set(self.group_indices.keys()))
        if len(groups) > 7 :
            colors = 'tab20b'
        else:
            colors = 'Dark2'
        n = max(len(groups) + 1, 2)
        self.cmap = cm.get_cmap(colors, n)
        self.group_colormap = {group: self.cmap((i+1)/n) for i,group in enumerate(groups)}
        self.std_color = self.cmap(0)

        # get standards
        self.stds = []
        for row in self.data_matrix.molecules.values():
            if row['std'] not in self.stds:
                self.stds.append(row['std'])

        # build combined outlier mask
        self.combined_outliers = np.zeros((self.n_samples, self.n_molecules), dtype=bool)
        for metric in self.data_matrix.outliers:
            self.combined_outliers |= self.data_matrix.outliers[metric]
        
        # setup per molcule metrics
        self.per_mol_metrics = {
                                '% Missing': self.data_matrix.missing.mean(axis=0) * 100,
                                "Mean Area": np.nanmean(self.data_matrix.data['norm_Area'], axis=0),
                                "SD Area": np.nanstd(self.data_matrix.data['norm_Area'], axis=0),
                                "% CV Area": (np.nanstd(self.data_matrix.data['norm_Area'], axis=0) / np.nanmean(self.data_matrix.data['norm_Area'], axis=0)) * 100,
                                "% Outliers": self.combined_outliers.mean(axis=0) * 100,
                                "Mean RT": np.nanmean(self.data_matrix.data['RT'], axis=0),
                                "SD RT": np.nanstd(self.data_matrix.data['RT'], axis=0),
                                "Mean S/N": np.nanmean(self.data_matrix.data['SN_Ratio'], axis=0),
                                "SD S/N": np.nanstd(self.data_matrix.data['SN_Ratio'], axis=0)
                                }

        # setup per sample metrics
        injection_order = np.zeros(self.n_samples)
        for row in self.data_matrix.samples.values():
            i = self.sample_map[row['sample_name']]
            injection_order[i] = row['injection_order']
        self.per_sample_metrics = {
                                    "% Missing": self.data_matrix.missing.mean(axis=1) * 100,
                                    "% Outliers": self.combined_outliers.mean(axis=1) * 100,
                                    "Injection Order": injection_order
                                    }

        # setup widgets
        self.data_table = QTableWidget()

        self.data_type_label = QLabel()
        self.data_type_dropdown = QComboBox()

        self.export_btn = QPushButton("Export to Excel")

        self.processing_label = QLabel("Loading...")

        self.initUI()

    def initUI(self):
        
        # data table columns
        self.h_header = ColoredHeader(Qt.Horizontal, self.data_table)
        self.data_table.setHorizontalHeader(self.h_header)
        col_count = self.n_molecules + len(list(self.per_sample_metrics)) + 1
        self.data_table.setColumnCount(col_count)
        col_labels = list(self.mol_map.keys())
        col_labels.append('')
        col_labels.extend(list(self.per_sample_metrics))
        for i,name in enumerate(col_labels):
            item = QTableWidgetItem(name)
            if name in self.stds:
                r,g,b,_ = [int(x*225) for x in self.std_color]
                item.setBackground(QColor(r,g,b,75))
            self.data_table.setHorizontalHeaderItem(i,item)
        # data table rows
        self.v_header = ColoredHeader(Qt.Vertical, self.data_table)
        self.data_table.setVerticalHeader(self.v_header)
        row_count = self.n_samples + len(list(self.per_mol_metrics)) + 1
        self.data_table.setRowCount(row_count)
        header_labels = list(self.sample_map.keys())
        header_labels.append('')
        header_labels.extend(self.per_mol_metrics)
        for i,name in enumerate(header_labels):
            item = QTableWidgetItem(name)
            if len(self.group_colormap) > 1:
                try:
                    group = self.group_map[name]
                    r,g,b,_ = [int(x*225) for x in self.group_colormap[group]]
                    item.setBackground(QColor(r,g,b,75))
                except:
                    pass
            self.data_table.setVerticalHeaderItem(i,item)

        # fill in per mol metrics
        for i,mol_metric in enumerate(self.per_mol_metrics):
            i += self.n_samples + 1
            for j,value in enumerate(self.per_mol_metrics[mol_metric]):
                self.data_table.setItem(i,j,QTableWidgetItem(self.format_val(value)))
        # fill in per sample metrics
        for j,sample_metric in enumerate(self.per_sample_metrics):
            j += self.n_molecules + 1
            for i,value in enumerate(self.per_sample_metrics[sample_metric]):
                self.data_table.setItem(i,j,QTableWidgetItem(self.format_val(value)))
        
        # darken empty cells for ease of viewing
        dark = QColor(0x23, 0x26, 0x29)
        i = self.n_samples
        for j in range(self.data_table.columnCount()):
            item = QTableWidgetItem()
            item.setBackground(dark)
            item.setFlags(Qt.NoItemFlags)
            self.data_table.setItem(i,j,item)
        j = self.n_molecules
        for i in range(self.data_table.rowCount()):
            item = QTableWidgetItem()
            item.setBackground(dark)
            item.setFlags(Qt.NoItemFlags)
            self.data_table.setItem(i,j,item)
        for i in range(self.n_samples, self.n_samples + len(self.per_mol_metrics) + 1):
            for j in range(self.n_molecules, self.n_molecules + len(self.per_sample_metrics) + 1):
                item = QTableWidgetItem()
                item.setBackground(dark)
                item.setFlags(Qt.NoItemFlags)
                self.data_table.setItem(i,j,item)

        # table policies
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.data_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.data_table.setItemDelegate(BackgroundDelegate(self.data_table))
        self.data_table.setFocusPolicy(Qt.ClickFocus)
        self.data_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_table.customContextMenuRequested.connect(self.on_cell_right_click)
        self.data_table.cellClicked.connect(self.on_cell_single_click)
        self.export_btn.clicked.connect(self.export_to_excel)

        # data type selection
        self.data_type_label.setText("Metric:")
        self.data_type_dropdown.addItems(self.data_metrics)
        self.data_type_dropdown.currentTextChanged.connect(self.data_type_changed)
        self.data_type_dropdown.setCurrentIndex(self.data_metrics.index('Area'))

        # setup table
        table_box = QGroupBox("Data")
        table_layout = QVBoxLayout()
        table_layout.addWidget(self.data_table)
        table_box.setLayout(table_layout)

        # setup side label
        info_box = QGroupBox("Info")

        outlier_section = QGroupBox("Outlier Metrics")
        self.outlier_layout = QVBoxLayout()
        none_label = QLabel("No Outliers")
        none_label.setStyleSheet("color: rgba(255,255,255,0.3); padding: 4px;")
        self.outlier_layout.addWidget(none_label)
        self.outlier_layout.addStretch()
        outlier_section.setLayout(self.outlier_layout)

        colors_section = QGroupBox("Group Colors")
        legend_layout = QVBoxLayout()
        for group,color in self.group_colormap.items():
            if len(self.group_colormap) > 1:
                label = QLabel(f"   {group}   ")
                r,g,b,_ = [int(x*225) for x in color]
                label.setStyleSheet(f"background-color: rgba({r},{g},{b},0.3); color: white; padding: 4px; border-radius: 3px;")
            else:
                label = QLabel("No Groups")
                label.setStyleSheet("color: rgba(255,255,255,0.3); padding: 4px;")
            legend_layout.addWidget(label)
        legend_layout.addStretch()
        colors_section.setLayout(legend_layout)

        info_layout = QVBoxLayout()
        info_layout.addWidget(outlier_section)
        info_layout.addWidget(colors_section)
        info_layout.addStretch()
        info_box.setLayout(info_layout)

        # splitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(table_box)
        splitter.addWidget(info_box)
        splitter.setStretchFactor(0,4)
        splitter.setStretchFactor(1,1)

        toolbar_layout = QHBoxLayout()
        toolbar_layout.addWidget(self.data_type_label)
        toolbar_layout.addWidget(self.data_type_dropdown)
        toolbar_layout.addWidget(self.processing_label)
        self.processing_label.setVisible(False)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.export_btn)

        layout = QVBoxLayout()
        layout.addLayout(toolbar_layout)
        layout.addWidget(splitter)
        self.setLayout(layout)

        self.data_type_changed()

    def data_type_changed(self):
        dtype = self.data_type_dropdown.currentText()
        self.processing_label.setVisible(True)

        self.worker = TableWorker(self.data_matrix, dtype, self.sample_map, self.mol_map,
                                  self.combined_outliers, self.format_val)
        
        self.worker.finished.connect(self.on_data_ready)
        self.worker.start()

        """
        self.data_table.setUpdatesEnabled(False)
        for sample in self.data_matrix.samples:
            i = self.sample_map[sample]
            for mol in self.data_matrix.molecules:
                j = self.mol_map[mol]
                item = QTableWidgetItem(self.format_val(self.data_matrix.data[dtype][i,j]))
                if self.combined_outliers[i,j]:
                    item.setBackground(QColor(180,0,0,75))
                self.data_table.setItem(i,j,item)
        self.data_table.setUpdatesEnabled(True)
        """

    def on_data_ready(self, results):
        self.data_table.setUpdatesEnabled(False)
        for i, j, value, is_outlier in results:
            item = QTableWidgetItem(value)
            if is_outlier:
                item.setBackground(QColor(180,0,0,75))
            self.data_table.setItem(i,j,item)
        self.data_table.setUpdatesEnabled(True)
        self.processing_label.setVisible(False)

    def on_cell_single_click(self, i, j):
        # handle invalid click locations
        if i >= self.n_samples or j >= self.n_molecules:
            # clear previous layout
            while self.outlier_layout.count():
                item = self.outlier_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            label = QLabel("No Outliers")
            label.setStyleSheet("color: rgba(255,255,255,0.4); padding: 4px;")
            self.outlier_layout.addWidget(label)
            self.outlier_layout.addStretch()
            return
        
        # clear previous layout
        while self.outlier_layout.count():
            item = self.outlier_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        out_types = [metric for metric in self.data_matrix.outliers
                     if self.data_matrix.outliers[metric][i,j]]
    
        if out_types:
            for metric in out_types:
                label = QLabel(f"   {metric}   ")
                label.setStyleSheet("background-color: rgb(180,60,60); color: white; padding: 4px; border-radius: 3px;")
                self.outlier_layout.addWidget(label)
        else:
            label = QLabel("No Outliers")
            label.setStyleSheet("color: rgba(255,255,255,0.4); padding: 4px;")
            self.outlier_layout.addWidget(label)
        
        self.outlier_layout.addStretch()
            
    def view_chrom(self, i, j):
        
        sample = self.inv_sample_map[i]
        molecule = self.inv_mol_map[j]

        if self.data_matrix.missing[i,j]:
            QMessageBox(self, "Error", f"No peak found for {molecule} in {sample}")
            return

        self.chrom_tab.navigate_to(sample, molecule)
        
        self.parent().parent().setCurrentWidget(self.chrom_tab)

    def on_cell_right_click(self, pos):

        index = self.data_table.indexAt(pos)
        i,j = index.row(), index.column()

        if i >= self.n_samples or j >= self.n_molecules:
            return
        
        menu = QMenu(self)
        view_chromatogram = menu.addAction("Chromatogram View")

        if menu.exec_(self.data_table.viewport().mapToGlobal(pos)) == view_chromatogram:
            self.view_chrom(i,j)

    def format_val(self, val):
        if val == 0:
            return "0"
        if abs(val) > 1000000 or abs(val) < 0.01:
            return f"{val:.4e}"
        return f"{val:.4f}"
    
    def export_to_excel(self):
        
        # set path and create workbook
        path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*xlsx)")
        if not path:
            return
        if not path.endswith('.xlsx'):
            path += '.xlsx'

        wb = Workbook()

        # aesthetic decisions
        dark_fill = self.make_fill("AA232629")
        grey_fill = self.make_fill("AABFBFBF")
        red_fill = self.make_fill("FFFF8080")
        bold_font = Font(bold=True)
        bottom_border = Border(bottom=Side(style='thin'))
        right_border = Border(right=Side(style='thin'))
        bottom_right_border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

        dtype = self.data_type_dropdown.currentText()
        raw_data = self.data_matrix.data[dtype]

        # first tab, metadata (samples and molecules tables) =====================================================
        ws1 = wb.active
        ws1.title = f"Metadata"
        
        # sample table
        sample_rows = list(self.data_matrix.samples.values())
        sample_cols = list(sample_rows[0].keys())
        
        # sample headers
        for j,col in enumerate(sample_cols):
            cell = ws1.cell(row=2, column=j+2, value=col)
            cell.fill = grey_fill
            cell.font = bold_font
            if j < len(sample_cols) - 1:
                cell.border = bottom_right_border
            else:
                cell.border = bottom_border

        # sample data
        for i,row in enumerate(sample_rows):
            for j,col in enumerate(sample_cols):
                ws1.cell(row=i+3, column=j+2, value=row[col])
                if j < len(sample_cols)-1:
                    ws1.cell(row=i+3, column=j+2).border = right_border

        # molecules table
        mol_rows = list(self.data_matrix.molecules.values())
        mol_cols = list(mol_rows[0].keys())
        
        # molecule headers
        for j,col in enumerate(mol_cols):
            cell = ws1.cell(row=2, column=len(sample_cols)+j+4, value=col)
            cell.fill = grey_fill
            cell.font = bold_font
            if j < len(mol_cols) - 1:
                cell.border = bottom_right_border
            else:
                cell.border = bottom_border

        # molecule data
        for i,row in enumerate(mol_rows):
            for j,col in enumerate(mol_cols):
                ws1.cell(row=i+3, column=len(sample_cols)+j+4, value=row[col])
                if j < len(mol_cols) - 1:
                    ws1.cell(row=i+3, column=len(sample_cols)+j+4).border = right_border

        # second tab, data table =================================================================================

        ws2 = wb.create_sheet(title=f"DataTable_{dtype}")
        
        # column labels
        for j in range(self.data_table.columnCount()):
            item = self.data_table.horizontalHeaderItem(j)
            cell = ws2.cell(row=2, column=j+3, value=item.text() if item else '')
            bg = item.background() if item else None
            if bg and bg.style() != Qt.NoBrush and bg.color().alpha() > 0:
                cell.fill = self.make_fill(self.qcolor_to_hex(bg.color()))
            elif j == self.n_molecules:
                cell.fill = dark_fill
            else:
                cell.fill = grey_fill
            cell.font = bold_font
            if j != self.n_molecules:
                cell.border = bottom_right_border if j < self.data_table.columnCount() - 1 else bottom_border
            
        # row labels
        for i in range(self.data_table.rowCount()):
            item = self.data_table.verticalHeaderItem(i)
            cell = ws2.cell(row=i+3, column=2, value=item.text() if item else '')
            bg = item.background() if item else None
            if bg and bg.style() != Qt.NoBrush and bg.color().alpha() > 0:
                cell.fill = self.make_fill(self.qcolor_to_hex(bg.color()))
            elif i == self.n_samples:
                cell.fill = dark_fill
            else:
                cell.fill = grey_fill
            cell.font = bold_font
            if i != self.n_samples:
                cell.border = right_border

        # top left
        ws2.cell(row=2, column=2).fill = grey_fill
        ws2.cell(row=2, column=2).border = bottom_right_border
        ws2.cell(row=1, column=2, value=f"Data:").font = bold_font
        ws2.cell(row=1, column=2).alignment = Alignment(horizontal='right')
        ws2.cell(row=1, column=3, value=dtype).font = bold_font

        # fill in data
        for i in range(self.data_table.rowCount()):
            for j in  range(self.data_table.columnCount()):
                cell = ws2.cell(row=i+3, column=j+3)

                # metric sections/empty space
                if i >= self.n_samples or j >= self.n_molecules:
                    table_item = self.data_table.item(i,j)
                    if table_item and table_item.text():
                        try:
                            cell.value = float(table_item.text())
                        except:
                            cell.value = table_item.text()
                        if j < self.data_table.columnCount() - 1:
                            cell.border = right_border
                    else:
                        cell.fill = dark_fill
                
                # main data
                else:
                    val = raw_data[i,j]
                    cell.value = None if np.isnan(val) else val
                    if j < self.data_table.columnCount() - 1 :
                        cell.border = right_border
                    if self.combined_outliers[i,j]:
                        cell.fill = red_fill

        # Third Tab, outliers ===================================================================================
        ws3 = wb.create_sheet(title="Outliers")

        # setup headers
        for j,header in enumerate(['sample_name', 'molecule_name', 'outlier_metrics']):
            cell = ws3.cell(row=2, column=j+2, value=header)
            cell.fill = grey_fill
            cell.font = bold_font
            if j < 2:
                cell.border = bottom_right_border
            else:
                cell.border = bottom_border

        # inv maps for index -> name/mol
        sample_names = {v:k for k,v in self.data_matrix.sample_map.items()}
        mol_names = {v:k for k,v in self.data_matrix.mol_map.items()}

        row_idx = 3
        for i in range(self.data_matrix.n_samples):
            for j in range(self.data_matrix.n_molecules):
                flagged = [metric for metric, matrix in self.data_matrix.outliers.items()
                           if matrix[i,j]]
                if flagged:
                    ws3.cell(row=row_idx, column=2, value=sample_names[i]).border = right_border
                    ws3.cell(row=row_idx, column=3, value=mol_names[j]).border = right_border
                    ws3.cell(row=row_idx, column=4, value=', '.join(flagged))
                    row_idx += 1

        self.autofit_columns(ws1)
        self.autofit_columns(ws2)
        self.autofit_columns(ws3)
        wb.save(path)

    def autofit_columns(self, ws, min_width = 8.43):
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value)) if  cell.value else 0) for cell in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width =max(max_len+2, min_width)

    def qcolor_to_hex(self, qcolor):
        return f"{qcolor.red():02X}{qcolor.green():02X}{qcolor.blue():02X}"

    def make_fill(self, hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

class BackgroundDelegate(QStyledItemDelegate):
    """
    Allows for highlighting of table cells 
    """
    def paint(self, painter, option, index):
        bg = index.data(Qt.BackgroundRole)
        if bg and isinstance(bg,QBrush) and bg.color().isValid():
            painter.fillRect(option.rect, bg)
        super().paint(painter,option,index)

class ColoredHeader(QHeaderView):
    """
    Allows for coloring of table headers
    """
    def __init__(self, orientation, table):
        super().__init__(orientation, table)
        self._table = table
        self.setStyleSheet("""
            QHeaderView {
                background-color: #232629
            }
            QHeaderView::section {
                background-color: transparent;
                color: rgba(255,255,255,0.7);
                text-transform: uppercase;
                padding: 0 24px;
                height: 36px;
                border-right: 1px solid #31363b;
                border-bottom: 1px solid #31363b;
            }
        """)

    def paintSection(self, painter, rect, logical_index):
        if self.orientation() == Qt.Horizontal:
            item = self._table.horizontalHeaderItem(logical_index)
        else:
            item = self._table.verticalHeaderItem(logical_index)

        bg = item.background() if item else None
        if bg and bg.style() !=Qt.NoBrush and bg.color().isValid() and bg.color().alpha() > 0:
            painter.fillRect(rect, bg)
        else:
            painter.fillRect(rect, QColor(0x23, 0x26, 0x29))

        super().paintSection(painter, rect, logical_index)

class TableWorker(QThread):
    finished = pyqtSignal(list)

    def __init__(self, data_matrix, dtype, sample_map, mol_map, outliers, format_val):
        super().__init__()
        self.data_matrix = data_matrix
        self.dtype = dtype
        self.sample_map = sample_map
        self.mol_map = mol_map
        self.outliers = outliers
        self.format_val = format_val

    def run(self):
        results = []
        for sample in self.data_matrix.samples:
            i = self.sample_map[sample]
            for mol in self.data_matrix.molecules:
                j = self.mol_map[mol]
                value = self.format_val(self.data_matrix.data[self.dtype][i,j])
                is_outlier = bool(self.outliers[i,j])
                results.append((i,j, value, is_outlier))
        self.finished.emit(results)