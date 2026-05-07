# region Imports

from pathlib import Path
import pandas as pd
import yaml, os
from openpyxl import load_workbook

# endregion

class ConfigLoader:
    """
    loads config.yaml and gives easy access to useful information
    """

    def __init__(self, config_file: Path):
        """
        Loads yaml file nd stores it as a dictionary as self.config
        params:
            config_file:            Path to config file, should be a relative path and it depends on where you run the script from in this project folder
        """
        self.config_path = Path(config_file)

        if not self.config_path.exists():
            raise FileNotFoundError(f"Config file at {config_file} not found")
        
        with open(self.config_path, "r") as f:
            self.config = yaml.safe_load(f)

    def get(self, *keys: str, default=None):
        """
        accesses nested values from config dict
        params:
            keys:                   list of keys in order of accessing for config structure
                example:            cfg.get("params","star","threads") returns number of threads tasked to star package
        """

        value = self.config
        
        # iterates over all keys given going into each key subsection at each iteartion
        for key in keys:

            # see if key is not in config

            # ensures key is a dict
            if not isinstance(value,dict):
                return default
            
            # resets value (dict) to the value under key, if key is a subdict name it reaturns that dict
            value = value.get(key, default)
        
        # return value queried for
        return value
    
    def get_path(self, *keys: str, base_path: Path = None, must_exist=False):
        """
        Returns Path object for the path specified in the config within the specified base_dir
        Params:
            keys:                   list of keys to go through to find the desired path
                example:            cfg.get_path("data_dirs", "raw") to get the path to the raw file data directory
            base_path:              Path object to concatenate the path we are retreiving, if we want to put subdir within another dir then we would make the dir the base_path
            must_exist:             default to False for directories that the script will create (such as those in data_dirs), true if it needs to exist before (like raw dir)
        """
        p = Path(self.get(*keys))

        if base_path and isinstance(base_path,Path):
            p = base_path / p

        if must_exist and not p.exists():
            raise FileNotFoundError(f"Path {p} not found for keys {keys}")
        
        return p

    def check_bools(self):

        # empty list to hold improperly formatted bools
        errors = []

        # list of fields that must be boolean
        bool_fields = {}

        # recursive function to enter parent dicts
        def recurse(value, path=""):
            # check each key value pair
            for k,v in value.items():
                # get string representation of current value being observed
                current_path = f"{path}.{k}" if path else k
                # if value is a dict then go another layer deeper
                if isinstance(v,dict):
                    recurse(v,current_path)
                # if not a dict, check if it is in bool_fields and if it is properly formatted as bool
                else:
                    # if imrpoperly formatted then append it to errors list
                    if k in bool_fields and not isinstance(v,bool):
                        errors.append(current_path)
        
        # run recursive method on loaded config dict
        recurse(self.config)
        
        # output error/all good message
        if errors:
            raise ValueError(
                f"Invalid boolean fields found in config.yaml, reformat to True/False:\n"+
                "\n".join(f" - {e}" for e in errors)
                )
        else:
            print("All boolean fields valid, continuing pipeline")

    def load_template(self):
        """
        Retreives the template file specified and returns the molecule, m/z, and rt lists
        Returns:
            molecule_data               dict of molecule information (for collection), key: list
            sample_data                 dict of sample information (for analysis), key: list

            all return values are matched by index, so the first molecule is idx 0 in all 3 lists
        """

        # get file values
        file_name = self.get("template_file")
        input_dir = self.get("input_dir")
        file = Path(input_dir) / file_name
        
        # read file (without header) and load moleucle, mz values, and retention times to lists
        df = pd.read_excel(file,skiprows=3)

        molecule_data = {
            "names": df["molecule"].dropna().to_list(),
            "mzs": df["mz"].dropna().to_list(),
            "rts": df["rt"].dropna().to_list()
        }

        sample_data = {
            "names": df["samples"].dropna().to_list(),
            "mouseIDs": df["mousID"].dropna().to_list(),
            "groups": df["group"].dropna().to_list(),
            "norms": df["norm"].dropna().to_list(),
            "inj_order": df["injection_order"].dropna().to_list()
        }

        return molecule_data, sample_data
    
    def generate_template(self):
        """
        generates a templeate xlsx file for inputting m/z and rt values
        """
        # get input dir
        input_dir = Path(self.get("input_dir"))
        out_dir = Path(self.get_path("results_dir"), input_dir)
        template = self.get("template_file")
        file = out_dir / template

        # get list of sample names from input dir
        names = sorted(
            p.stem
            for p in input_dir.iterdir()
            if p.is_dir() and p.suffix == ".D"
        )

        # generate sample table df
        sample_df = pd.DataFrame({
            "samples": names,
            "mouseID": ['' for _ in names],
            "group": ['' for _ in names],
            "norm": ['' for _ in names],
            "injection_order": ['' for _ in names]
        })

        # generate data df 
        data_df = pd.DataFrame(columns=["molecule","mz","rt","standard"])

        # generate headers
        header1 = "Template file for gcms automatic peak picking/integration, please ONLY fill in appropriate values and feel free to leave case/control empty if need be"
        header2 = "group = grouping for samples (ie case/control), norm = normalization factor, molecule = id of this moleucue, mz = ion to measure, rt = peak retention time, standard = name of standard to apply to that sample"

        # add sample/data dfs to excel file
        with pd.ExcelWriter(file, engine="openpyxl") as writer:

            sample_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=1
            )

            data_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=7
            )

        # add headers
        wb = load_workbook(file)
        ws = wb.active

        ws["A1"] = header1
        ws["A2"] = header2

        wb.save(file)
