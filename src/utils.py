"""

miscelaneous utility functions for use across the project

"""

# region Imports

import subprocess,json,shutil,sys,re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

# logging
import logging
logger = logging.getLogger(__name__)

# endregion

def log_subprocess(result: subprocess.CompletedProcess, log_dir: Path, id: str):
    """
    Collects logs produced by python subprocess and puts them in a text file for easy viewing
    Params:
        result:                     subprocess.CompletedProcess object that can be logged, result of running subprocess
        log_dir:                    path to the directory to store logs in
        id:                         identifier string for this process being logged
    """
    # get timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # make sure log_dir exist
    log_dir.mkdir(parents=True,exist_ok=True)

    # path to log file
    log_file = log_dir / "subprocess_log.jsonl"

    # dict of values to store
    data = {
        "id": id,
        "log_ts": timestamp,
        "stdout": result.stdout,
        "stderr": result.stderr
    }

    with open(log_file, "a") as f:
        f.write(json.dumps(data) + "\n")

def generate_template(input_dir: Path, file: Path = "template.xlsx"):
    """
    generates a templeate xlsx file for iputting m/z and rt values
    Params:
        input_dir                       input directory for files to be analyzed
        file                            name/location of template file
    """

    header1 = "Template file for gcms automatic peak picking/integration, please ONLY fill in appropriate values and feel free to leave case/control empty if need be"
    header2 = "molecule = id of this moleucue, mz = ion to measure, rt = peak retention time case/control = list of sample names in each group (as they appear in input file names)"

    df = pd.DataFrame(columns=["molecule","mz","rt","case","control"])
    df.to_excel(Path(input_dir) / file, index=False,startrow=3,startcol=1)

    wb = load_workbook(file)
    ws = wb.active

    ws["A1"] = header1
    ws["A2"] = header2

    wb.save(file)

def log_timestamp(input_dir: Path, label: str, ts: datetime):
    """
    logs the timestamp of an operation
    Params:
        input_dir                       iput directory where we will be storing the results of the ts log
        label                           label for this part of the ts log
        ts                              datetime object of the timestamp for this operation
    """
    log_file = Path(input_dir) / "timestamp_log.jsonl"

    data = {
        label: ts.isoformat()
    }

    with open(log_file, "a") as f:
        f.write(json.dumps(data) + "\n")

def delete_file(file: Path):
        """ 
        Deletes a specifed file
        Params:
            file                Path to the file to be deleted
        """
        if file.exists():
            file.unlink()

def delete_directory(dir_path: Path):
    """
    Deletes a specifed directory and all incldued files
    Params:
       dir_path                 Path to the directory to delete
    """
    try:
        shutil.rmtree(dir_path)
    except FileNotFoundError as e:
        print(f"File:\n{dir_path}\nnot found\nError:\n{e}")
    except PermissionError as p:
        print(f"Permission error deleting:\n{dir_path}\nError:\n{p}")

def get_app_dir():
    """
    Gets the directroy for the appliation, if frozen (windows app) then put databases right
    next to installation .exe, if scripting then put databases in root

    Returns
    -------
    path to the application directory (root or .exe)
    """
    if getattr(sys,'frozen',False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent.parent

def get_run_dir(proj_name, run_name):
    return get_app_dir() / 'databases' / 'projects' / proj_name / run_name

def get_proj_db(proj_name):
    return get_app_dir() / 'databases' / 'projects' / proj_name / f"{proj_name}.db"

def get_proj_dir(proj_name):
    return get_app_dir() / 'databases' / 'projects' / proj_name

def get_run_cfg_path(proj_name, run_name):
    return get_run_dir(proj_name, run_name) / 'config.yaml'

def configure_run_logging(run_dir: Path):
    """
    Redirects the root logger's file output to run_dir/log_name, replacing any
    previously attached FileHandler. Safe to call again later (e.g. switching runs).
    """
    run_dir = Path(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)

    root_logger = logging.getLogger()

    for handler in root_logger.handlers[:]:
        if isinstance(handler, logging.FileHandler):
            root_logger.removeHandler(handler)
            handler.close()
            
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
    file_handler = logging.FileHandler(run_dir / "run.log")
    file_handler.setFormatter(formatter)
    root_logger.addHandler(file_handler)

def sanitize_name(name: str):
    """
    Replaces problematic characaters and spaces with "_"
    """
    name = name.strip()
    name = re.sub(r'[\\/:*?"<>|]','_',name)
    name = name.replace(" ", "_")
    return name

def find_abandoned_runs(proj_dir: Path):
    if not proj_dir.is_dir():
        logger.warning(f"No directory found at {proj_dir}")
        return
    
    abandoned = []
    for entry in proj_dir.iterdir():
        if not entry.is_dir():
            continue
        if not any(entry.glob("*.h5")):
            abandoned.append(entry)

    return abandoned
    