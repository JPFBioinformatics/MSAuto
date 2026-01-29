# region Imports

import sys, math
from pathlib import Path
from collections import defaultdict
from matplotlib.backends.backend_pdf import PdfPages
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors

# location of pipeline root dir
root_dir = Path(__file__).resolve().parent.parent
# tell python to look here for modules
sys.path.insert(0, str(root_dir))

from src.config_loader import ConfigLoader

# endregion

class ReportGenerator:
    """
    Class that will generate a report based on a list of identified peaks
    """

    def __init__(self, cfg: ConfigLoader, peaks: dict = None):
        """
        Params:
            cfg                     loaded config for this run
            peaks                   dict of peak values for each sample (sample: list of peak dicts)
        """
        self.cfg = cfg                          # config file
        self.peaks = peaks                      # peak dictionary
        self.matrix = None                      # raw data matrix
        self.value_map = None                   # column map (molecule name: index) for the raw matrix
        self.norm_matrix = None                 # matrix that has been normalized to internal standards/normalization factors
        self.norm_value_map = None              # column map for the normalized matrix
        self.clean_matrix = None                # norm_matrix where columns with high NaN ratios dropped and remaining NaN values imputed
        self.clean_value_map = None             # column map for the clean matrix
        self.sample_map = None                  # row map (sample name: index)
        self.std_cols = None                    # dictionary of standard name: standard column from raw matrix
        self.group_index_map = None             # dictionary of group_name: list of column index values for that group
        self.output_flags = None                # flag dictionary for coloring the output data tables
        self.outliers = None                    # dictionary of outliers for outlier flagging
        self.nan_ratios = None                  # dictionary of NaN ratios for each column
        self.temlate_dict = None                # dictionary produced by reading the template file

    # region Matrix Generation

    def save_peaks(self, peaks: dict):
        """
        Saves a peak dict as self.peaks if not specified when object is creatd
        Params:
            peaks                   dict of sample: peak list where each peak is a dict for the full set of samples for data analysis
        """
        self.peaks = peaks
    
    def normalize_matrix(self):
        """
        Generates two new matrices, a normalized matrix (divide all values by internal std and normalization factor) as well as a clean matrix
        that starts as the normalized matrix, then has columns with large NaN ratios removed and the remaining NaN values are replaced via imputation
        """

        # get data matrix
        raw_matrix = self.matrix
        if raw_matrix is None:
            raise ValueError("No data matrix detected, please run generate_matrix before normalizing")

        # load template values
        self.read_template()
        template_dict = self.template_dict

        # normalize to internal standard and norm_factor
        standards = template_dict['standards']
        molecules = template_dict['molecules']
        if len(standards) > 0:
            norm_matrix, norm_map = self.normalize_istd(raw_matrix,standards,molecules)
        else:
            raise ValueError("Please specify standards in template file")
        norm = template_dict['norm_factors']
        samples = template_dict['samples']
        if len(norm) > 0:
            norm_matrix = self.normalize_amount(norm_matrix,norm,samples)

        # save normalized values
        self.norm_matrix = norm_matrix
        self.norm_value_map = norm_map

        # clean data (remove NaN values and impute)
        clean_matrix, clean_map = self.drop_sparse_features(norm_matrix,norm_map)
        clean_matrix = self.impute_nans(clean_matrix)

        # save cleaned data
        self.clean_matrix = clean_matrix
        self.clean_value_map = clean_map

        # flag outliers
        self.flag_outliers()

    def read_template(self):
        """
        reads a given template file and returns the relevant data, lists are matched by index
        Returns:
            data                        dict of loaded values
        """
        # load config and template file
        cfg = self.cfg
        template_name = cfg.get("template_file")
        file = Path(cfg.get("input_dir")) / template_name

        # read template file and 
        df = pd.read_excel(file,skiprows=3)

        # grab values
        data = {}
        data['samples'] = self.get_col(df,'samples')
        data['groups'] = self.get_col(df,'group')
        data['norm_factors'] = self.get_col(df,'norm')
        data['standards'] = self.get_col(df,'standard')
        data['molecules'] = self.get_col(df,'molecule')
        data['mz'] = self.get_col(df,'mz')
        data['rt'] = self.get_col(df,'rt')

        
        # build group index map
        group_index_map = {}
        for sample,group in zip(data['samples'],data['groups']):
            sample_i = self.sample_map[sample]
            if group not in group_index_map:
                group_index_map[group] = [(sample_i)]
            else:
                group_index_map[group].append(sample_i)
            
        self.group_index_map = group_index_map
        self.template_dict = data
    
    def normalize_istd(self, matrix: np.ndarray, standards: list, molecules: list):
        """
        Normalizes a matrix to its internal standard(s)
        Params:
            matrix                          matrix to nomralize
            standards                       list of standard names to normalize by, one standard per molecule measured (lots of reapeaded strings)
            molecules                       list of molecules names, index matched to standards list (from template)
        Returns:
            matrix
        """
        # get the feature map and data matrix
        feature_map = self.value_map
        data = matrix.copy()
        std_set = set(standards)
        feature_to_std = dict(zip(molecules,standards))

        # copy standard columns for later use
        std_cols = {}
        std_idxs = []
        for standard in std_set:
            std_i = feature_map[standard]
            values = data[:,std_i].copy()
            if np.any(values <= 0) or np.any(~np.isfinite(values)):
                raise ValueError(f"{standard} contains zero/NaN/ivalid value(s), cannot normalize")
            std_cols[standard] = values
            std_idxs.append(std_i)

        # save standard columns (std_name:column) to object
        self.std_cols = std_cols
        std_idxs_set = set(std_idxs)
        std_idxs_list = sorted(std_idxs_set)

        # divide each value in the row by its corrosponding istd value (each sample has its own istd value)
        for name,col_i in feature_map.items():
            if name in std_set:
                continue
            std_name = feature_to_std[name]
            std_vals = std_cols[std_name]
            data[:,col_i] /= std_vals

        # remove istd columns from matrix
        new_data = np.delete(data,std_idxs_list,axis=1)

        # reprocess value map to match this new matrix
        inv_map = {idx:name for name,idx in feature_map.items()}
        new_map = {}
        new_i = 0
        for old_i in range (data.shape[1]):
            if old_i in std_idxs:
                continue
            name = inv_map[old_i]
            new_map[name] = new_i
            new_i += 1

        return new_data,new_map

    def normalize_amount(self, matrix: np.ndarray, norm: list, samples: list):
        """
        Normalizes the matrix to a suplied norm list, representing mass/volume/protein amount etc... of sample
        Params:
            matrix                          matrix to normalize
            norm                            list of values to normalize by (float)
            samples                         list of samples in the run 
        """
        # generate a map of sample to normalization value
        sample_to_norm = dict(zip(samples,norm))

        # copy matrix for normalization
        data = matrix.copy()
        sample_map = self.sample_map

        # loop over all sample rows and normalize to the norm value
        for sample,norm_val in sample_to_norm.items():
            row_i = sample_map[sample]
            data[row_i,:] /= norm_val

        return data

    def get_col(self, df: pd.DataFrame, col_name: str):
        """
        Gets column list from an excel file loaded as a df
        Params:
            df                              df to parse
            col_name                        name of column you want
        Returns:
            values                          list of values read
        """
        return df[col_name].dropna().to_list() if col_name in df.columns else []

    def drop_sparse_features(self, matrix: np.ndarray, col_map: dict, max_nan_ratio: float = 0.5):
        """
        Drops features from the matrix (before full normalization) that have a lot of np.nan values (missing data)
        Params:
            matrix                          the matrix you want to process, matrix MUST have one-hot encoded groups at the end
            col_map                         map of col_name: col_idx for the input matrix
            max_nan_ratio                   ratio of number nan_values / len(column) for rejection (per group)
        """

        # get group information
        group_index_map = self.group_index_map

        # get data information
        data = matrix.copy()
        num_features = data.shape[1]

        # calculate nan ratio for each column
        nan_ratios = {}
        for group_name,rows in group_index_map.items():
            sub = data[rows,:]
            nan_ratios[group_name] = np.isnan(sub).mean(axis=0)
        self.nan_ratios = nan_ratios
        
        # generate mask for columns to keep, keep column if ANY group has valid nan ratio
        keep = np.zeros(data.shape[1],dtype=bool)
        for group_name in group_index_map:
            group_keep = nan_ratios[group_name] <= max_nan_ratio
            keep |= group_keep

        # filter the data matrix
        filtered_data = data[:,keep]

        # rebuild column map to reflect filtered_data with nan columns removed
        inv_map = {idx:name for name,idx in col_map.items() if idx < num_features}
        new_map = {}
        new_i = 0
        for old_i in range(num_features):
            if keep[old_i]:
                new_map[inv_map[old_i]] = new_i
                new_i += 1

        return filtered_data, new_map

    def impute_nans(self, matrix: np.ndarray, factor: float = 0.5):
        """
        Replaces NaN values in matrix with an imputed value based on min measured value * factor
        Params:
            matrix                      matrix to impute
            factor                      Factor by which to multiply the lowest measured value to produce the value to replace NaN's for in that column
        """
        # get array of min values per column in matrix (ignoring NaN values)
        mins = np.nanmin(matrix,axis=0)

        # caluclate imputation value for each col
        imp_vals = factor * mins

        # replace nan values with calculated value
        nan_mask = np.isnan(matrix)
        matrix[nan_mask] = imp_vals[np.where(nan_mask)[1]]

        return matrix

    def generate_matrix(self, molecules: list):
        """
        Takes abundance/area data and stores it in a matrix as well as generating maps for samples and molecules
        matrix is organized with rows = samples and columns = molecules
        method will also reformat peaks dict so that it is simply sample_name: peak area list for each peak
        the molecule and sample maps are used to decode this
        Params:
            molecules                       list of molecules to be represented with a row of the matrix
        """

        # get peak data and generate maps
        if self.peaks:
            peaks = self.peaks
        else:
            raise ValueError("No peak dict found")

        # generate sample map
        sample_names = list(peaks.keys())
        num_samples = len(sample_names)
        sample_map = {}
        for idx,name in enumerate(sample_names):
            sample_map[name] = idx
        num_peaks = len(peaks[sample_names[0]])

        # generate molecule map
        molecule_map = {}
        for idx,molecule in enumerate(molecules):
            molecule_map[molecule] = idx

        # save peak data in a 2d numpy array
        data = np.zeros((num_samples,num_peaks))

        # lists to hold the i,j positions of cells we want to color
        # red for invalid rt, orange for flat top peaks, yellow for small peaks and blue for extra wide/overloaded peaks
        red = []
        yellow = []
        orange = []
        blue = []
        normal = []
        
        # iterate over all peaks, store data in matrix and store extra info where needed
        for i,sample in enumerate(sample_names):
            for j,peak in enumerate(peaks[sample]):

                if peak['width_flag'] == "small":
                    info = {
                        "coords": (i,j),
                        "issue": "small peak"
                    }
                    data[i,j] = peak["area"]
                    yellow.append(info)

                if peak['width_flag'] == "overloaded":
                    info = {
                        "coords": (i,j),
                        "issue": "wide peak"
                    }
                    data[i,j] = peak["area"]
                    blue.append(info)

                if peak["flat_top"]:
                    info ={
                        "coords": (i,j),
                        "issue": "flat top"
                    }
                    data[i,j] = peak["area"]
                    orange.append(info)

                if not peak["rt_valid"]:
                    data[i,j] = np.nan
                    info = {
                        "coords": (i,j),
                        "issue": peak["rt_diff"]
                    }
                    red.append(info)

                else:
                    info = {
                        "coords": (i,j),
                        "issue": "None"
                    }
                    data[i,j] = peak['area']
                    normal.append(info)
        
        # save flags dict for coloring of output files
        flags = {
            "red": red,
            "orange": orange,
            "yellow": yellow,
            "blue": blue,
            "normal": normal
        }
        self.output_flags = flags

        # save data to object
        self.matrix = data
        self.sample_map = sample_map
        self.value_map = molecule_map

    def stack_new_col(self, matrix: np.ndarray, values: np.ndarray, label: str):
        """
        Method to add a new column to the data array, used to insert case/control, sex etc.. labels
        Params:
            values                      valid list to add as a new column
            label                       label to place new column under in value map
        """
        # ensure values and matrix are compatible size
        num_rows = matrix.shape[0]
        num_cols = matrix.shape[1]
        if len(values) != num_rows:
            raise ValueError(f"Incorrect column lenght, {len(values)} rows cannot be added to a matrix with {num_rows} rows")
        
        # if compatible then continue
        new_col = np.array(values).reshape(-1,1)
        matrix = np.hstack((matrix,new_col))

        # add to value map
        self.value_map[label] = num_cols

        # save new matrix
        self.matrix = matrix

    # endregion

    # region QC Calculations
    """ PCA Note
    PCA calculations are in the pca_plots function under PDF/Excel Generation region
    """

    def qc_data(self):
        """
        Generates data dicts for QC plotting
        """

        if self.peaks is None:
            raise ValueError("No peak data availabe for QC plots")
        if self.outliers is None:
            raise ValueError("No outlier data available for QC plots")
        
        # get metric data
        sample_data = {}
        total_data = {
            "flat_top": [],
            "width_flag": [],
            "widths": [],
            "rt_diffs": [],
            "rt_valid": [],
            "symmetry": [],
            "sym_valid": []
        }
        
        for sample,values in self.peaks.items():

            # dict to hold data for this sample
            data = {
                "flat_top": [],
                "width_flag": [],
                "widths": [],
                "rt_diffs": [],
                "rt_valid": [],
                "symmetry": [],
                "sym_valid": []
            }

            for peak in values:
       
                # if its an empty peak then skip it
                if peak['width_flag'] == 'N/A':
                    continue

                # extract lists of values for analysis
                ft = peak["flat_top"]
                wf = peak["width_flag"]
                w = peak["right_bound"] - peak["left_bound"]
                rt = peak["rt_diff"]
                rv = peak["rt_valid"]
                sym = peak["bound_symmetry"]
                sv = peak["symmetry_valid"]

                # per-sample
                data["flat_top"].append(ft)
                data["width_flag"].append(wf)
                data["widths"].append(w)
                data["rt_diffs"].append(rt)
                data["rt_valid"].append(rv)
                data["symmetry"].append(sym)
                data["sym_valid"].append(sv)

                # total
                total_data["flat_top"].append(ft)
                total_data["width_flag"].append(wf)
                total_data["widths"].append(w)
                total_data["rt_diffs"].append(rt)
                total_data["rt_valid"].append(rv)
                total_data["symmetry"].append(sym)
                total_data["sym_valid"].append(sv)

            # calculate pct outliers per sample
            if self.pct_outliers is not None:
                pct_int = self.pct_outliers['sample']['intensity'].get(sample,0)
                pct_rt = self.pct_outliers['sample']['rt_diff'].get(sample,0)
                data['% Intensity Outliers'] = pct_int
                data['% RT Outliers'] = pct_rt
            else:
                data['% Intensity Outliers'] = 0
                data['% RT Outliers'] = 0

            sample_data[sample] = data
        
        # calculate total pct outliers
        if self.pct_outliers is not None:
            total_int = list(self.pct_outliers['sample']['intensity'].values())
            total_rt = list(self.pct_outliers['sample']['rt_diff'].values())
            total_data['% Intensity Outliers'] = np.mean(total_int)
            total_data['% RT Outliers'] = np.mean(total_rt)
        else:
            total_data['% Intensity Outliers'] = 0
            total_data['% RT Outliers'] = 0
        return total_data, sample_data

    def compute_stats(self, flags_dict: dict):
        """
        Helper method to compute QC stats from total or sample data dicts
        Params:
            flags_dict                      dict that contains flag info for calculation
        """

        # calcualte other metrics
        num_peaks = len(flags_dict["width_flag"])
        flat_top_pct = 100 * sum(flags_dict["flat_top"]) / num_peaks
        width_counts = Counter(flags_dict["width_flag"])
        width_pct = {k: 100*v/num_peaks for k,v in width_counts.items()}
        avg_width = np.mean(flags_dict["widths"])
        err_width = np.std(flags_dict["widths"]) / np.sqrt(num_peaks)
        rt_valid_pct = 100 * (sum(flags_dict["rt_valid"]) / num_peaks)
        avg_rt = np.mean(flags_dict["rt_diffs"])
        err_rt = np.std(flags_dict["rt_diffs"]) / np.sqrt(num_peaks)
        sym_valid_pct = 100 * (1 - np.mean(flags_dict["sym_valid"]))
        avg_sym = np.mean(flags_dict["symmetry"])
        err_sym = np.std(flags_dict["symmetry"]) / np.sqrt(num_peaks)
        rt_outs = flags_dict['% RT Outliers']
        int_outs = flags_dict['% Intensity Outliers']

        return {
            r"% FlatTop": flat_top_pct,
            "% Narrow": width_pct.get("small",0),
            "% Ideal": width_pct.get("ideal",0),
            "% Normal": width_pct.get("normal",0),
            "% Overloaded": width_pct.get("overloaded",0),
            "\u03BC Width": avg_width,
            "\u03C3 Width": err_width,
            "% RT Valid": rt_valid_pct,
            "\u03BC \u0394RT": avg_rt,
            "\u03C3 \u0394RT": err_rt,
            "% Valid Sym": sym_valid_pct,
            "\u03BC Sym": avg_sym,
            "\u03C3 Sym": err_sym,
            "% RT Outliers": rt_outs,
            "% Intensity Outliers" : int_outs
            }
    
    def identify_outliers(self, matrix: np.ndarray, threshold: float = 3.5):
        """
        Helper method that takes an input matrix and calculates outliers in each column
        Params:
            matrix                  the matrix you want to find outliers in
            threshold               threshold for mod_z scores that denotes an outlier
        """
        # structures to hold data about which columns were skipped
        skipped_cols = set()
        eligable_cols_by_sample = defaultdict(set)
        outliers = []

        # inverse maps to map row/col to sample/feature
        inv_sample = {idx:name for name,idx in self.sample_map.items()}
        inv_value = {idx:name for name,idx in self.value_map.items()}
        
        # iterate over all columns of the matrix
        for col_i,col in enumerate(matrix.T):
            # mask to remove nan values
            mask = ~np.isnan(col)
            if mask.sum() < 3:
                skipped_cols.add(col_i)
                continue
            nonzero_vals = col[mask]
            
            # calculate median and MAD
            median = np.median(nonzero_vals)
            mad = np.median(np.abs(nonzero_vals-median))

            # skip column if there is no variation
            if mad == 0:
                skipped_cols.add(col_i)
                continue
            # skip column if it is sparse
            if mask.sum() / len(col) > 0.5:
                skipped_cols.add(col_i)
                continue

            # grab row indices for eligable column
            row_indices = np.where(mask)[0]

            # mark eligability per sample
            for row_i in row_indices:
                sample = inv_sample[row_i]
                eligable_cols_by_sample[sample].add(col_i)
            
            # compute modz
            mod_z = 0.6745 * (nonzero_vals - median) / mad

            # flag outliers
            for row_i,z in zip(row_indices,mod_z):
                if abs(z) > threshold:
                    outliers.append({
                        'row': row_i,
                        'col': col_i,
                        'sample' : inv_sample[row_i],
                        'molecule' : inv_value[col_i],
                        'mod_z' : float(z)
                    })

        return outliers,eligable_cols_by_sample

    def flag_intensity_outliers(self, threshold: float = 3.5):
        """
        Flags outlier values columnwise in the matrix, allowing for finding of raw abundance values that stick out from the distribution expected for that molecule
        uses a modified MAD (median absolute deviation) based z-score function
        Params:
            theshold                threshold value for samples to be flagged
        """
        # grab data matrix
        data = self.norm_matrix.copy()
        """
        # get inverse maps so we can map idx:name of mol/sample
        inv_sample = {idx:name for name,idx in self.sample_map.items()}
        inv_value = {idx:name for name,idx in self.value_map.items()}
        outliers = []
        
        # find outliers in each column
        skipped_cols = set()
        eligalbe_cols_by_sample = defaultdict(set())
        for col_i, col in enumerate(data.T):
            
            # generate mask to remove nan values
            mask = ~np.isnan(col)
            if mask.sum() < 3:
                skipped_cols.add(col_i)
                continue
            nonzero_col = col[mask]
            median = np.median(nonzero_col)
            mad = np.median(np.abs(nonzero_col-median))

            # if no variation then skip this column
            if mad == 0:
                skipped_cols.add(col_i)
                continue
            # skip sparse features
            if mask.sum() / len(col) < 0.5:
                skipped_cols.add(col_i)
                continue
            
            # calculate modified z scor
            mod_z = 0.6745 * (nonzero_col-median) / mad

            # find row index values of outliers
            row_indices = np.where(mask)[0]
            for row_i, val in zip(row_indices,mod_z):
                if abs(val) > threshold:
                    values = {
                        'row': row_i,
                        'col': col_i,
                        'sample': inv_sample[row_i],
                        'molecule': inv_value[col_i],
                        'mod_z': float(val)
                    }
                    outliers.append(values)

        return outliers, data.shape[1] - skipped_cols
        """
        # identify outliers
        outliers, eligable_counts = self.identify_outliers(data)
        return outliers, eligable_counts

    def flag_rt_outliers(self, threshold: float = 3.5):
        """
        Finds outliers in the RT of collcted peaks
        Params:
            threshold                       float value to threshold the z-scored variance on (above is an outlier)
        """

        # get peaks dict and sample list
        peaks = self.peaks
        samples = list(peaks.keys())
        num_molecules = len(self.value_map)
        num_samples = len(samples)
        #inv_sample = {idx:name for name,idx in self.sample_map.items()}
        #inv_value = {idx:name for name,idx in self.value_map.items()}

        data = np.full((num_samples,num_molecules), np.nan)
        """
        outliers = []
        
        # iterate through all peaks for each sample, placing rt_diff in the correct place in the matrix
        for sample in samples:
            row_idx = self.sample_map[sample]
            for entry in peaks[sample]:
                col_idx = self.value_map[entry["molecule"]]
                rt_matrix[row_idx,col_idx] = entry["rt_diff"]

        # now detect outliers in each column and save rt_outliers = value_name :  sample_name for an outlier
        skipped_cols = 0
        for col_i, col in enumerate(rt_matrix.T):

            # generate mask to remove nan values
            mask = ~np.isnan(col)
            if mask.sum() < 3:
                continue
            nonzero_col = col[mask]
            median = np.median(nonzero_col)
            mad = np.median(np.abs(nonzero_col-median))

            # if no variation then skip column
            if mad == 0:
                skipped_cols += 1
                continue
            # don't calculate outliers for spase features
            if mask.sum() / len(col) < 0.5:
                skipped_cols += 1
                continue
            
            # calculate mod z score
            mod_z = 0.6745 * (nonzero_col-median) / mad

            # collect data
            row_indices = np.where(mask)[0]
            for row_i,val in zip(row_indices,mod_z):
                if abs(val) > threshold:
                    values = {
                        'row': row_i,
                        'col': col_i,
                        'sample': inv_sample[row_i],
                        'molecule': inv_value[col_i],
                        'mod_z': float(val),
                    }
                    outliers.append(values)

        return outliers
        """
        # identify outliers
        outliers, eligable_counts = self.identify_outliers(data)
        return outliers,eligable_counts

    def flag_outliers(self):
        """
        Function that will flag all rt and intensity (types) and calculates percent outliers per sample and per molecule
        """
        # inverse value maps
        inv_val = {idx:name for name,idx in self.value_map.items()}
        inv_norm_val = {idx:name for name,idx in self.norm_value_map.items()}

        # find outliers in intensity and rt
        int_outliers,int_counts = self.flag_intensity_outliers()
        rt_outliers,rt_counts = self.flag_rt_outliers()
        int_counts_mol = {mol:0 for mol in self.norm_value_map.keys()}
        rt_counts_mol = {mol:0 for mol in self.value_map.keys()}
        
        # initialize dicts for counting outliers/sample
        pct_rt_outliers = {sample: 0 for sample in self.sample_map.keys()}
        pct_rt_outliers_mol = {mol: 0 for mol in self.value_map.keys()}
        pct_int_outliers = {sample: 0 for sample in self.sample_map.keys()}
        pct_int_outliers_mol = {mol: 0 for mol in self.norm_value_map.keys()}

        # count outliers per sample
        for out in int_outliers:
            pct_int_outliers[out['sample']] += 1
            pct_int_outliers_mol[out['molecule']] += 1
        for out in rt_outliers:
            pct_rt_outliers[out['sample']] += 1
            pct_rt_outliers_mol[out['molecule']] += 1
        # calculate % outlier per sample
        for sample,count in pct_int_outliers.items():
            denom = len(int_counts.get(sample,set()))
            pct_int_outliers[sample] = 100 * count / denom if denom > 0 else np.nan
        for sample,count in pct_rt_outliers.items():
            denom = len(rt_counts.get(sample,set()))
            pct_rt_outliers[sample] = 100 * count / denom if denom > 0 else np.nan

        # count eligable counts per molecule
        for sample,cols in int_counts.items():
            for col_i in cols:
                mol = inv_norm_val[col_i]
                int_counts_mol[mol] += 1
        for sample,cols in rt_counts.items():
            for col_i in cols:
                mol = inv_val[col_i]
                rt_counts_mol[mol] += 1
        # compute % outliers per molecule
        for mol,count in pct_int_outliers_mol.items():
            denom = int_counts_mol.get(mol,0)
            pct_int_outliers_mol[mol] = 100 * count / denom if denom > 0 else np.nan
        for mol,count in pct_rt_outliers_mol.items():
            denom = rt_counts_mol.get(mol,0)
            pct_rt_outliers_mol[mol] = 100 * count / denom if denom > 0 else np.nan

        # save values
        self.outliers = {
            'intensity' : int_outliers,
            'rt_diff' : rt_outliers
        }
        self.pct_outliers = {
            'sample' : {
                'intensity' : pct_int_outliers,
                'rt_diff' : pct_rt_outliers
            },
            'molecule' : {
                'intensity' : pct_int_outliers_mol,
                'rt_diff' : pct_rt_outliers_mol
            }
        }

    # endregion

    # region PDF/Excel Generation

    def graph_nans(self, pdf):
        """
        Generates a 3 panel report for nan values, boxplot, histogram, and heatmap
        """

        # put nan ratios into a df
        df = pd.DataFrame(self.nan_ratios)
        df.index.name = "feature"

        # creat figure
        fig,axes = plt.subplots(1,3,figsize=(18,5))

        # generate boxplots per group
        sns.boxenplot(data=df,ax=axes[0])
        axes[0].set_title("Missingness by Group")
        axes[0].set_ylabel("Nan Ratio")
        axes[0].set_xlabel("Group")

        # generate histogram of missingness (group agnostic)
        all_nan = df.values.flatten()
        sns.histplot(all_nan,bins=20,kde=False,ax=axes[1])
        axes[1].set_title("Overall Missingness Distribution")
        axes[1].set_xlabel("NaN Ratio")
        axes[1].set_ylabel("Feature Count")

        # generate heatmap
        sns.heatmap(df.T, cmap="viridis",cbar=True,ax=axes[2])
        axes[2].set_title("MIssingness Heatmap (Group x Features)")
        axes[2].set_xlable("Feature Index")
        axes[2].set_ylable("Group")

        # save figure
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def std_qc_plots(self, pdf):
        """
        Generates a box and whisker polot for standard values, naming samples that are outliers in standard amount
        Params:
            pdf                         pdf file to save figure to
        """
        # get name and number of standards
        std_names = list(self.std_cols.keys())
        num_stds = len(std_names)

        # generate figure
        fig,axes = plt.subplots(1,num_stds,figsize=(num_stds*2,6), squeeze=False)

        for i,std_name in enumerate(std_names):

            # geet values for this std
            values = self.std_cols[std_name]

            # plot
            ax = axes[0,i]
            bp = ax.boxplot(values, vert=True, showfliers=True)
            ax.set_title(f"Standard {std_name}")
            ax.set_ylabel("Abundance")
            ax.set_xticks([])

            # annotate fliers
            fliers = bp["fliers"][0].get_ydata()
            for y in fliers:
                idx = (values == y).nonzero()[0]
                for i in idx:
                    ax.annotate(
                        i,
                        xy=(1,y),
                        xytext=(1.05,y),
                        fontsize=8,
                        va="center"
                    )

        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def qc_df(self, pdf, sample_data: dict, total_data: dict, rows_per_page: int = 25):
        """
        Generates a pandas dataframe and then plots it as a data table
        Params:
            sample_data                     sample:data where data is another dict
            total_data                      same as sample_data but summed for the entire run, not per sample
        """

        # add sample map page
        self.add_sample_map_page(pdf)

        # compute stats per sample
        rows = []
        for sample,data in sample_data.items():
            row = self.compute_stats(data)
            row["sample"] = self.sample_map[sample]
            rows.append(row)

        # compute stats for total
        total_row = self.compute_stats(total_data)
        total_row["sample"] = "Total"
        rows.append(total_row)

        # create df
        df_qc = pd.DataFrame(rows)
        df_qc = df_qc[[
            "sample",
            r"% FlatTop",
            "% Narrow",
            "% Ideal",
            "% Normal",
            "% Overloaded",
            "\u03BC Width",
            "\u03C3 Width",
            "% RT Valid",
            "\u03BC \u0394RT",
            "\u03C3 \u0394RT",
            "% Valid Sym",
            "\u03BC Sym",
            "\u03C3 Sym",
            "% RT Outliers",
            "% Intensity Outliers"
        ]]
        # add table to QC PDF
        self.add_table_to_pdf(pdf,df_qc,"QC Summary Table",rows_per_page)
        
        return df_qc

    def generate_template(self):
        """
        generates a templeate xlsx file for inputting m/z and rt values
        """
        # get input dir
        cfg = self.cfg
        input_dir = Path(cfg.get("input_dir"))
        out_dir = Path(cfg.get_path("results_dir"), input_dir)
        template = cfg.get("template_file")
        file = out_dir / template

        # get list of sample names from input dir
        names = sorted(
            p.stem
            for p in input_dir.iterdir()
            if p.is_dir() and p.suffix == ".D"
        )

        # generate sample table df
        sample_df = pd.DataFrame({
            "samples": names,
            "group": ['' for _ in names],
            "norm": ['' for _ in names],
        })

        # generate data df 
        data_df = pd.DataFrame(columns=["molecule","mz","rt","standard"])

        # generate headers
        header1 = "Template file for gcms automatic peak picking/integration, please ONLY fill in appropriate values and feel free to leave case/control empty if need be"
        header2 = "group = grouping for samples (ie case/control), norm = normalization factor, molecule = id of this moleucue, mz = ion to measure, rt = peak retention time, standard = name of standard to apply to that sample"

        # add sample/data dfs to excel file
        with pd.ExcelWriter(file, engine="openpyxl") as writer:

            sample_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=1
            )

            data_df.to_excel(
                writer,
                index=False,
                startrow=3,
                startcol=5
            )
        
        # add headers
        wb = load_workbook(file)
        ws = wb.active

        ws["A1"] = header1
        ws["A2"] = header2

        wb.save(file)

    def generate_report(self, num_pcs: int = 5):
        """
        Generates a single pdf report for the entire run
        """
        cfg = self.cfg
        input_dir = Path(cfg.get("input_dir"))
        res = cfg.get("results_dir")
        out_dir = input_dir / res
        name = cfg.get("run_name")
        out_file = Path(out_dir) / f"{name}_report.pdf"

        if self.norm_matrix is None:
            raise ValueError("Normalize Matrix before generating report")

        with PdfPages(out_file) as pdf:
            self.add_metadata_pages(pdf)
            self.std_qc_plots(pdf=pdf)
            total,per_sample = self.qc_data()
            df = self.qc_df(pdf,per_sample,total)
            self.plot_outliers(pdf)
            self.plot_qc_total(pdf,total)
            self.plot_qc_per_sample(pdf,df)
            self.plot_qc_per_molecule(pdf)
            self.pca_plots(pdf=pdf,num_comps=num_pcs)

    def add_metadata_pages(self, pdf):
        """
        adds metadata pages to the report pdf showing group membership/collection information
        Params:
            pdf                         pdf to save the page to
        """
        metadata = self.template_dict
        num_samples = len(metadata['samples'])
        num_molecules = len(metadata['molecules'])

        # sample table
        sample_df = pd.DataFrame({
            'Sample': metadata['samples'],
            'Group': metadata['groups'] if metadata['groups'] else ['' for _ in range(num_samples)],
            'Norm Factor': metadata['norm_factors'] if metadata['norm_factors'] else ['' for _ in range(num_samples)]
        })

        # molecule table
        mol_df = pd.DataFrame({
            'Molecule': metadata['molecules'],
            'Ion': metadata['mz'],
            'Retention Time': metadata['rt'],
            'Standard': metadata['standards'] if metadata['standards'] else ['' for _ in range(num_molecules)]
        })

        # add pages to pdf
        self.add_table_to_pdf(pdf,sample_df,"Sample Metadata")
        self.add_table_to_pdf(pdf,mol_df,"Molecule Metadata")

    def add_sample_map_page(self, pdf):
        """
        Adds a PDF page mapping sample names to sample indices
        Params:
            pdf                 pdf to save this figure to
        """

        # Convert sample_map to DataFrame
        data = sorted(self.sample_map.items(), key=lambda x: x[1])
        df = pd.DataFrame(data, columns=["Sample Name", "Sample Index"])

        # Create figure
        fig, ax = plt.subplots(figsize=(8, 10))
        ax.axis("off")

        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            loc="center",
            cellLoc="center"
        )

        # Formatting
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.0, 1.2)

        fig.suptitle(
            "Sample Index Mapping",
            fontsize=14,
            y=0.92
        )

        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def add_feature_map_page(self,pdf):
        """
        Adds a page giving feature name to feature number map
        Params:
            pdf                             pdf file to save to
        """
        # Convert sample_map to DataFrame
        data = sorted(self.value_map.items(), key=lambda x: x[1])
        df = pd.DataFrame(data, columns=["Sample Name", "Sample Index"])

        # Create figure
        fig, ax = plt.subplots(figsize=(8, 10))
        ax.axis("off")

        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            loc="center",
            cellLoc="center"
        )

        # Formatting
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.0, 1.2)

        fig.suptitle(
            "Sample Index Mapping",
            fontsize=14,
            y=0.92
        )

        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)
  
    def plot_qc_total(self, pdf, total_data: dict, sym_bin: float = 0.1, rt_bin: float = 0.01, width_bin: float = 1.0):
        """
        Plots QC metrics for the full set of samples
        Params:
            total_data                          dataframe of calculated values
            sym/rt/width_bin                    Bin size for the histograms for these metrics
        """

        cfg = self.cfg
        sym_threshold = cfg.get("endpoint_threshold")
        rt_threshold = cfg.get("rt_threshold")
        width_threshold = cfg.get("width_threshold")

        # find maxes and mins for binning
        width_min = min(total_data["widths"])
        width_max = max(total_data["widths"])
        sym_min = min(total_data["symmetry"])
        sym_max = max(total_data["symmetry"])
        rt_min = min(total_data["rt_diffs"])
        rt_max = max(total_data["rt_diffs"])

        # bin
        sym_bins = np.arange(
            sym_min,
            sym_max + sym_bin,
            sym_bin
        )
        rt_bins = np.arange(
            rt_min,
            rt_max + rt_bin,
            rt_bin
        )
        width_bins = np.arange(
            width_min,
            width_max + width_bin,
            width_bin
        )

        # genreate histogram figure
        fig,axes = plt.subplots(3,2,figsize=(8,10))
        fig.suptitle("Total Run QC",fontsize=14)
        fig.tight_layout(rect=[0,0,1,0.95])

        # symmetry histogram
        symmetry_values = total_data["symmetry"]
        axes[0,0].hist(symmetry_values,bins=sym_bins)
        axes[0,0].set_xlim(sym_min,sym_max)
        axes[0,0].set_title("Bound Symmetry Distribution (Right - Left)")
        axes[0,0].set_xlabel("Bound Symmetry")
        axes[0,0].set_ylabel("Count")
        axes[0,0].axvline(
            sym_threshold,
            linestyle="--",
            linewidth=2,
            label="Symmetry Threshold"
        )
        axes[0,0].axvline(
            -1*sym_threshold,
            linestyle="--",
            linewidth=2,
            label="Symmetry Threshold"
        )
        axes[0,0].legend()
        
        # symmetry box and whisker plot
        axes[0,1].boxplot(symmetry_values,vert=True,showfliers=True)
        axes[0,1].set_title("Bound Symmetry Plot (Right - Left)")
        axes[0,1].set_ylabel("Symmetry Value")

        # rt difference histogram
        rt_diffs = total_data["rt_diffs"]
        axes[1,0].hist(rt_diffs,bins=rt_bins)
        axes[1,0].set_xlim(rt_min,rt_max)
        axes[1,0].set_title("RT Difference Distribution")
        axes[1,0].set_xlabel("RT Difference (min)")
        axes[1,0].set_ylabel("Count")
        axes[1,0].axvline(
            rt_threshold,
            linestyle="--",
            linewidth=2,
            label="RT Threshold"
        )
        axes[1,0].axvline(
            -1*rt_threshold,
            linestyle="--",
            linewidth=2,
            label="RT Threshold"
        )
        axes[1,0].legend()

        # rt diffs box and whisker plot
        axes[1,1].boxplot(rt_diffs,vert=True,showfliers=True)
        axes[1,1].set_title("RT Difference Plot")
        axes[1,1].set_ylabel("RT Difference")
       

        # width histogram
        widths = total_data["widths"]
        axes[2,0].hist(widths,bins=width_bins)
        axes[2,0].set_xlim(width_min,width_max)
        axes[2,0].set_title("Width Distribution")
        axes[2,0].set_xlabel("Width (scans)")
        axes[2,0].set_ylabel("Count")
        axes[2,0].axvline(
            width_threshold,
            linestyle="--",
            linewidth=2,
            label="Width Threshold"
        )
        axes[2,0].legend()

        # width box and whisker plot
        widths = total_data["widths"]
        axes[2,1].boxplot(widths,vert=True,showfliers=True)
        axes[2,1].set_title("Width Plot")
        axes[2,1].set_ylabel("Width")
        
        # save figure
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def plot_qc_per_sample(self, pdf, df_qc: pd.DataFrame):
        """
        Generates box and whisker plots for per seample metrics, such as average width etc.. and plots them, labelling outliers
        """

        # metrics to plot
        metrics = [
            r"% FlatTop",
            "% RT Valid",
            "% Valid Sym",
            "\u03BC Width",
            "\u03C3 Width",
            "\u03BC \u0394RT",
            "\u03C3 \u0394RT",
            "\u03BC Sym",
            "\u03C3 Sym",
            "% Intensity Outliers",
            "% RT Outliers"
        ]

        # generate plot
        fig,axes = plt.subplots(4,3, figsize=(10,10))
        fig.suptitle("Per-Sample QC", fontsize=14)
        
        # generate figures per metric
        for idx,metric in enumerate(metrics):
            
            # position the figure
            row = idx // 3
            col = idx % 3
            ax = axes[row,col]

            # error if the df does not have the relevant data
            if metric not in df_qc.columns:
                raise ValueError(f"Metric {metric} not found in samples dataframe:\n{df_qc.columns}")
            
            # handle the rest of the plots
            else:

                # get y and x values
                metric_values = df_qc[metric].values
                samples = df_qc["sample"].tolist()

                # handle metrics with no data (happens with # Outliers as well as others)
                if len(metric_values) == 0:
                    metric_values = [0]

                # creat subplot
                bp = ax.boxplot(metric_values, vert=True, showfliers=True)
                ax.set_title(metric)
                ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

                # annotate fliers
                fliers = bp["fliers"][0].get_ydata()
                for y in fliers:
                    idx = [j for j,val in enumerate(metric_values) if val == y]
                    for i in idx:
                        ax.annotate(
                            samples[i],
                            xy=(1,y),
                            xytext=(1.05,y),
                            fontsize=5,
                            va="center"
                        )

        pdf.savefig(fig)
        plt.close(fig)

    def plot_qc_per_molecule(self,pdf):
        """
        generates boxplots for per molecule metrics of interest, all calculated from normalized matrix EXCEPT the % rt outliers
        %RT outliers, %Intensity Outliers, medain intensity, % missingness, and Coefficeint of variation per molecule
        """

        # check that data is present
        if self.norm_matrix is None:
            raise ValueError("Normalized matrix is required for per-molecule QC plotting")

        # get per molecule data
        molecules = list(self.norm_value_map.keys())
        data = self.norm_matrix
        n_mol = data.shape[1]

        # % outliers per molecule
        all_molecules = list(self.value_map.keys())
        pct_rt_out = [self.pct_outliers['molecule']['rt_diff'].get(mol,0) for mol in all_molecules]
        pct_int_out = [self.pct_outliers['molecule']['intensity'].get(mol,0) for mol in molecules]


        # calculate median intensity, missingness, and cv values
        med_int = []
        missing = []
        cv = []
        for col in range(n_mol):
            vals = data[:,col]
            med_int.append(np.nanmedian(vals))
            missing.append(100*np.isnan(vals).sum()/len(vals))
            mean = np.nanmean(vals)
            std = np.nanstd(vals)
            cv.append(std/mean if mean != 0 else 0)

            # metrics to plot
        metrics = [
            ('% RT Outliers',pct_rt_out),
            ('% Intensity Outliers',pct_int_out),
            ('Median Intensity',med_int),
            ('% Missing Values',missing),
            ('Coefficient of Variation',cv)
        ]

        # build the figure
        fig,axes = plt.subplots(3,2,figsize=(10,10))
        axes = axes.flatten()
        fig.suptitle("Per-Molecule QC",fontsize=16)

        for met_i,(title,values) in enumerate(metrics):
            ax = axes[met_i]
            bp = ax.boxplot(values,vert=True,showfliers=True)
            ax.set_title(title)
            ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

            # annotate fliers
            fliers = bp["fliers"][0].get_ydata()
            for y in fliers:
                values_arr = np.array(values)
                idx = [j for j,val in enumerate(values_arr) if val == y]
                for i in idx:
                    if title == '% RT Outliers':
                        label = all_molecules[i]
                    else:
                        label = molecules[i]
                    ax.annotate(
                        label,
                        xy=(1,y),
                        xytext=(1.05,y),
                        fontsize=5,
                        va="center"
                    )

        pdf.savefig(fig)
        plt.close(fig)

    def pca_plots(self, pdf, num_comps: int = 2):
        """
        Generates a PCA plot for the data stored in this object (self.matrix) and saves output to report pdf
        Params:
            pdf                             pdf file to save figure to
            num_comps                       the number of PCA components to include in plot
        """
        # raise error if normalized matrix does not exist
        if self.clean_matrix is None:
            raise ValueError("No normalized matrix available for PCA")

        # get matrix and remove the one-hot encoded groups if needed
        data = self.clean_matrix.copy()

        # calculate value for imputing nan 
        min_nonzero = np.nanmin(data,axis=0)
        data = np.where(np.isnan(data), min_nonzero/2, data)

        # grab threshold values
        cfg = self.cfg
        var_threshold = cfg.get("variance_threshold")
        pca_var = cfg.get("pca_var")

        # convert to ppm
        data = (data / data.sum(axis=1,keepdims=True)) * 1e6

        # log transform
        data = np.log1p(data)

        # filter out low varaince features
        feature_var = data.var(axis=0)
        keep = feature_var > float(var_threshold)
        data = data[:,keep]

        # crate new feature map to reflect filtered matrix
        start_map = self.clean_value_map
        feature_names = [name for name,_ in sorted(start_map.items(), key=lambda x:x[1])]
        filtered_names = np.array(feature_names)[keep]
        filtered_map = {name:i for i,name in enumerate(filtered_names)}

        # z-score transform
        data = StandardScaler().fit_transform(data)

        # calculate PCA and explained variance
        pca = PCA(n_components=num_comps)
        cg = False
        if self.group_index_map:
            cg = True
        scores = pca.fit_transform(data)
        variance = pca.explained_variance_ratio_

        # plot variance bar graph
        fig,ax=plt.subplots(figsize=(7,5), constrained_layout=True)

        ax.bar(range(1,num_comps+1), variance, color="skyblue", edgecolor = 'k', label="Explained Variance")
        ax.set_xlabel("Principal Component")
        ax.set_ylabel("Explained Variance Ratio")
        ax.set_title("PCA explained Variance")
        ax.set_xticks(range(1,num_comps+1))
        ax.legend()

        # save figure
        pdf.savefig(fig)
        plt.close(fig)

        # determine how many pcs to plot (those that sum to account for 80% of varaince by default)
        cumulative_var = np.cumsum(variance)
        num_to_plot = np.searchsorted(cumulative_var, pca_var) + 1
        num_to_plot = min(num_to_plot,scores.shape[1])

        # plot relevant pca combinations
        for i in range(num_to_plot - 1):
            for j in range(i+1,num_to_plot):

                pc_pair = scores[:,[i,j]]
                
                # plot pca scatter plot
                self.plot_pca(pc_pair,num_pcx=i+1, num_pcy=j+1, pdf=pdf, color_groups=cg)

                # get top loadings for this pca pair
                x_loadings = self.get_top_features(pca, filtered_map, i)
                y_loadings = self.get_top_features(pca, filtered_map, j)

                # build loadings dataframe
                loadings = pd.DataFrame({
                    f'PC{i+1} Feature': list(x_loadings.keys()),
                    f'PC{i+1} Loadings': list(x_loadings.values()),
                    f'PC{j+1} Feature': list(y_loadings.keys()),
                    f'PC{j+1} Loadings': list(y_loadings.values())
                })

                # add table to PDF
                self.add_table_to_pdf(pdf,loadings,f"Top Features: PC{j+1} vs PC{i+1}",rows_per_page=10)

    def plot_pca(self, pc_scores, num_pcx: int, num_pcy: int, pdf, color_groups: bool = True, color_molecule: str = None):
        """
        Generates a PCA plot for a given two principal components, helper method for calculate_pca
        Params:
            pc_scores                       scores extracted from the PCA object
            num_pcx/num_pcy                 numbers of the pcs being graphed (for labels)
            pdf                             pdf file to save figures to
            color_groups                    bool True if you want to color by group false if you don't want to
            color_molecule                  name of a molecule that you want to color by abundance of
        """

        fig,ax = plt.subplots(figsize=(6,6))

        # color samples based on gruoup (explicit, not continuous)
        if color_groups:

            # assign colors to groups
            unique_groups = sorted(self.group_index_map.keys())
            colors = plt.cm.tab10.colors
            group_color_map = {g:colors[i%10] for i,g in enumerate(unique_groups)}

            # plot PCA with each group colored
            for group_name, row_list in self.group_index_map.items():
                idx = np.array(row_list)
                ax.scatter(
                    pc_scores[idx,0],
                    pc_scores[idx,1],
                    label=group_name,
                    color=group_color_map[group_name],
                    alpha=0.8
                    )

            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
            ax.set_title(f"PC{num_pcy} vs PC{num_pcx}")
            ax.legend()

        # color samples based on a specifed value
        elif color_molecule:

            # get column from matrix that corrosponds to the molecule to color by
            col_idx = self.molecule_map[color_molecule]
            values = self.matrix[:, col_idx]

            # normalize and generate color maps
            norm = mcolors.Normalize(vmin=np.min(values), vmax=np.max(values))
            cmap = mcolors.LinearSegmentedColormap.from_list("custom", ["blue","red"])

            # plot
            sc = ax.scatter(pc_scores[:,0], pc_scores[:,1], c=values, cmap=cmap, norm=norm, alpha=0.8)
            fig.colorbar(sc).set_label(f"{color_molecule} Abundance")
            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
            ax.set_title(f"{num_pcy} vs {num_pcx}")
        
        # do not color samples if not specified
        else:
            ax.scatter(pc_scores[:,0], pc_scores[:,1], alpha=0.8)
            ax.set_title(f"{num_pcy} vs {num_pcx}")
            ax.set_xlabel(f"PC {num_pcx}")
            ax.set_ylabel(f"PC {num_pcy}")
        
        # save figure   
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def get_top_features(self, pca, feature_map: dict, pc_index: int, n=10):
        """
        finds the top n features that contribute most to this pc of your pca
        Params:
            pca                             output from PCA run on you rmatrix
            feature_map                     map of feature_name:column_index for the cleaned/filtered matrix used for PCA
            pc_index                        index value of the pc you want to get features from
            n                               number of feature/loading pairs to retrieve
        Returns:
            top_loadings                    dict of col_index : loading score for the top n loadings for this pc (pc_index of pca)
        """
        # generate inverse feature map for easy access to feature names
        inv_feat_map = {i:name for name,i in feature_map.items()}

        # get top n loadings
        loadings = np.abs(pca.components_[pc_index])
        top_idx = np.argsort(loadings)[-n:][::1]

        # place in a dict
        top_loadings = {}
        for idx in top_idx:
            top_loadings[inv_feat_map[idx]] = loadings[idx]

        return top_loadings

    def write_to_excel(self):
        """
        Generates the output excel file
        """
        # grab config values
        cfg = self.cfg
        template = cfg.get('template_file')
        input_dir = Path(cfg.get("input_dir"))
        template_location = input_dir / template
        name = cfg.get("run_name")
        results = cfg.get("results_dir")
        results_dir = input_dir / results
        results_dir.mkdir(parents=True,exist_ok=True)
        out_file = Path(results_dir) / f"{name}.xlsx"

        # grab all the metadata
        metadata = self.template_dict
        num_samples = len(metadata['samples'])
        num_molecules = len(metadata['molecules'])

        # set up dict to inform coloring of cells
        fills = {
            "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "orange": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
            "yellow": PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),
            "blue": PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
        }
        # sort sample and molecule maps to ensure correct ordering
        samples_ordered = [sample for sample, _ in sorted(self.sample_map.items(), key=lambda x: x[1])]
        molecules_ordered = [mol for mol, _ in sorted(self.value_map.items(), key=lambda x: x[1])]
        if self.norm_value_map is not None:
            norm_molecules_ordered = [mol for mol, _ in sorted(self.norm_value_map.items(), key=lambda x: x[1])]
        if self.clean_value_map is not None:
            clean_molecules_ordered = [mol for mol, _ in sorted(self.clean_value_map.items(), key=lambda x: x[1])]

        # generate excel file
        with pd.ExcelWriter(out_file,engine="openpyxl") as writer:

            # ===== Metadata Tab =====
            # sample table
            sample_df = pd.DataFrame({
                'Sample': metadata['samples'],
                'Group': metadata['groups'] if metadata['groups'] else ['' for _ in range(num_samples)],
                'Norm Factor': metadata['norm_factors'] if metadata['norm_factors'] else ['' for _ in range(num_samples)]
            })
            sample_df.to_excel(writer, sheet_name="Metadata", startrow=4, startcol=2, index=False)

            # molecule table
            mol_df = pd.DataFrame({
                'Molecule': metadata['molecules'],
                'Ion': metadata['mz'],
                'Retention Time': metadata['rt'],
                'Standard': metadata['standards'] if metadata['standards'] else ['' for _ in range(num_molecules)]
            })
            mol_df.to_excel(writer, sheet_name="Metadata", startrow=4, startcol=6, index=False)

            # add tables/header to tab
            ws_meta = writer.sheets["Metadata"]
            ws_meta['B2'] = f"Metadata for reference only, to change analysis edit template sheet at: {template_location}"

            # ===== Raw Data Tab =====
            df_raw = pd.DataFrame(self.matrix, index=samples_ordered, columns=molecules_ordered)
            df_raw.to_excel(writer, sheet_name="Raw", index=True)

            # color the sheet
            ws = writer.sheets["Raw"]

            # genreate legend
            ws.insert_rows(1,8)
            ws["A1"] = "Flag Key:"
            ws["A2"] = "Red = Invalid RT"
            ws["A3"] = "Orange = Flat-Top Peak"
            ws["A4"] = "Yellow = Small Peak"
            ws["A5"] = "Blue = Overloaded/Wide Peak"
            ws["A6"] = "Dark Red Text = RT Outlier"
            ws["A7"] = "Purple Text = Intensity Outlier"

            for color, entries in self.output_flags.items():
                # skip coloring "normal" peaks
                if color not in fills:
                    continue
                    
                # get fill color
                fill = fills[color]
                    
                # color cells
                for entry in entries:
                    i,j = entry["coords"]
                        
                    excel_row = i+10
                    excel_col = j+2

                    ws.cell(row=excel_row, column=excel_col).fill = fill

            # add text color to RT outliers
            outlier_font = {
                'rt_diff': Font(color="C00000", bold=True),
                'intensity': Font(color="7030A0", bold=True)
            }
            if self.outliers:
                for out_type,value_list in self.outliers.items():
                    for entry in value_list:
                        i = int(entry['row'])
                        j = int(entry['col'])
                        row_i = i+10
                        col_i = j+2

                        cell = ws.cell(row=row_i,column=col_i)
                        cell.font = outlier_font[out_type]

            # ===== Normalized Matrix =====
            if self.norm_matrix is None:
                raise ValueError(f"Normailze matrix before returning normalized data")
                
            df_norm = pd.DataFrame(self.norm_matrix, index=samples_ordered, columns=norm_molecules_ordered)
            df_norm.to_excel(writer, sheet_name="Normalized", index=True)

            # ===== Cleaned Matrix ======
            if self.clean_matrix is None:
                raise ValueError(f"Generate Clean matrix before outputting")
            
            df_clean = pd.DataFrame(self.clean_matrix, index=samples_ordered, columns=clean_molecules_ordered)
            df_clean.to_excel(writer, sheet_name="Cleaned")

    def plot_outliers(self, pdf, rows_per_page: int = 25):
        """
        Adds tables of outliers for RT and Intensity
        Params
            pdf                         PDF file to save this figure to
            rows_per_page               Number of rows to add to the table per page
        """

        rt_outs = self.outliers['rt_diff']
        int_outs = self.outliers['intensity']

        if len(rt_outs) > 0:
            rt_df = pd.DataFrame(rt_outs)
            self.add_table_to_pdf(pdf,rt_df,"RT Outliers",rows_per_page)

        if len(int_outs) > 0:
            int_df = pd.DataFrame(int_outs)
            self.add_table_to_pdf(pdf,int_df,"Intensity Outliers",rows_per_page)

    def add_table_to_pdf(self, pdf, df: pd.DataFrame, title: str, rows_per_page: int = 25):
        """
        takes a dataframe and plots it as a table on a supplied pdf
        Params:
            pdf                         PDF to save to
            df                          dataframe to convert to a table
            title                       title of the table
            rows_per_page               how many rows to include per page of PDF
        """

        num_pages = math.ceil(len(df) / rows_per_page)
        for page in range(num_pages):

            # find rows for this page
            start = page * rows_per_page
            end = start + rows_per_page
            df_chunk = df.iloc[start:end]
            formatted = df_chunk.copy()
            for col in formatted.columns:
                try:
                    formatted[col] = formatted[col].apply(lambda x: f"{float(x):.3g}")
                except:
                    continue

            # generate figure
            fig,ax = plt.subplots(figsize=(8.5,11))
            fig.tight_layout()
            ax.axis("off")

            # draw table
            table = ax.table(
                cellText = formatted.values,
                colLabels = formatted.columns,
                loc = "center",
                cellLoc = "center"
            )

            # basic table settings
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0,1.5)
                
            # format column titles for readability
            for key,cell in table.get_celld().items():
                row,col = key
                if row == 0:
                    cell.visible_edges = "LR"
                    text = cell.get_text()
                    text.set_rotation(60)
                    text.set_verticalalignment("bottom")
                    text.set_fontsize(8)
                    text.set_fontweight("bold")
                    text.set_x(0.3)

            # add title
            if num_pages > 1:
                title += f" (Page {page+1} of {num_pages})"
            fig.suptitle(title, fontsize=14, y=0.92)

            pdf.savefig(fig)
            plt.close(fig)

    # endregion
